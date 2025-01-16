import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def pushToExcel(inDataFrame):

##ws = book.worksheets[0]
##for cell in ws["C"]:
##    if cell.value is None:
##        print cell.row
##        break
##else:
##    print cell.row + 1


    df = inDataFrame
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Data'

    wb.title = 'Data Export'

    for cell in sheet['A']:
        if cell.value is None:
            print(cell.row)
            nextCell = cell.row + 1
            break
    else:
        print(cell.row + 1)
    
    for row in df.iterrows():
        sheet.append(row[1].tolist())

    for row in sheet.rows:
##        max_length = 0
##        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                 pass

                
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

        sheet.row_dimensions[1].height = 12

     



        filename = r'C:\Users\brian\Documents\Spreadsheets\Project\Porsche.xlsx'
        wb.save(filename)


fileList = ["GE_MR_OEM_4242234_9X3M2Y.xlsm","GE_MR_OEM_4242234_8X3M2Y.xlsm","GE_MR_OEM_4242234_7X3M2Y.xlsm"]

customerInfoList = []
refNumList = []
serviceTagList = []
techName = ['Brian','Brian','Brian']

for item in fileList:

    #gets customer name from file
    customerInfoList.append(item[-35:-20])
    # gets refnum from file
    refNumList.append(item[-19:-13])
    # gets service tag from file
    serviceTagList.append(item[-11:-5])
    # need to go into the file to grab the tech name
    

data = []
data.append(customerInfoList)
data.append(refNumList)
data.append(serviceTagList)
data.append(techName)

print(data)
data = pd.DataFrame(data).transpose()
data.columns =['Customer Name','Ref#','Service','Tech']

print(data)
#pushToExcel(data)

